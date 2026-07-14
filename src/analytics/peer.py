import os
import sqlite3
from pathlib import Path
from typing import List

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Styles for Excel export
PERCENTILE_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # >= 75th
PERCENTILE_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # 25th to 75th
PERCENTILE_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # <= 25th
BENCHMARK_GOLD = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")   # Gold row
BORDER_THIN = Border(
    left=Side(style='thin', color='D3D3D3'),
    right=Side(style='thin', color='D3D3D3'),
    top=Side(style='thin', color='D3D3D3'),
    bottom=Side(style='thin', color='D3D3D3')
)

class PeerAnalyzer:
    def __init__(self, db_path: str = "data/db/nifty100.db", peer_xlsx_path: str = "data/raw/peer_groups.xlsx"):
        self.db_path = db_path
        self.peer_xlsx_path = peer_xlsx_path
        self.peer_df = self.load_peer_groups()

    def load_peer_groups(self) -> pd.DataFrame:
        """Loads peer groups mapping from excel sheet."""
        if not os.path.exists(self.peer_xlsx_path):
            raise FileNotFoundError(f"Peer groups file not found at: {self.peer_xlsx_path}")
        return pd.read_excel(self.peer_xlsx_path)

    def get_company_peer_group(self, ticker: str) -> str:
        """Returns the peer group name for a ticker, or fallback message."""
        ticker_clean = ticker.strip().upper()
        match = self.peer_df[self.peer_df["ticker"].str.strip().str.upper() == ticker_clean]
        if match.empty:
            return "No peer group assigned"
        return match.iloc[0]["group_name"]

    def calculate_percentiles_and_save(self):
        """Computes PERCENT_RANK for 10 metrics within each peer group and year, then populates DB."""
        conn = sqlite3.connect(self.db_path)

        query = """
            SELECT 
                fr.*,
                c.name as company_name,
                c.sector_name as broad_sector
            FROM financial_ratios fr
            JOIN companies c ON fr.ticker = c.ticker
        """
        df = pd.read_sql_query(query, conn)

        metrics_to_rank = [
            "roe", "return_on_capital_employed_pct", "net_profit_margin_pct",
            "debt_to_equity", "free_cash_flow_cr", "pat_cagr_5yr",
            "revenue_cagr_5yr", "eps_cagr_5yr", "interest_coverage", "asset_turnover"
        ]

        df["interest_coverage_ranked"] = df.apply(
            lambda r: float('inf') if r.icr_label == "Debt Free" else (r.interest_coverage or 0.0), axis=1
        )

        peer_dict = self.peer_df.set_index("ticker")["group_name"].to_dict()
        df["peer_group"] = df["ticker"].map(peer_dict)

        df_mapped = df[df["peer_group"].notna()].copy()

        cursor = conn.cursor()
        cursor.execute("DROP TABLE IF EXISTS peer_percentiles;")
        cursor.execute("""
            CREATE TABLE peer_percentiles (
                company_id TEXT,
                peer_group_name TEXT,
                metric TEXT,
                value REAL,
                percentile_rank REAL,
                year INTEGER,
                PRIMARY KEY (company_id, peer_group_name, metric, year)
            );
        """)
        conn.commit()

        records = []
        years = df_mapped["year"].unique()
        groups = df_mapped["peer_group"].unique()

        for year in years:
            year_mask = df_mapped["year"] == year
            df_year = df_mapped[year_mask]

            for group in groups:
                group_mask = df_year["peer_group"] == group
                df_group = df_year[group_mask]

                n = len(df_group)
                if n == 0:
                    continue

                for metric in metrics_to_rank:
                    col_to_use = "interest_coverage_ranked" if metric == "interest_coverage" else metric
                    vals = df_group[col_to_use].copy()
                    vals = vals.fillna(0.0)

                    ranks = vals.rank(method='min')

                    if n <= 1:
                        pct_ranks = pd.Series(1.0, index=vals.index)
                    else:
                        pct_ranks = (ranks - 1.0) / (n - 1.0)

                    if metric == "debt_to_equity":
                        pct_ranks = 1.0 - pct_ranks

                    for idx, pct in pct_ranks.items():
                        ticker = df_group.loc[idx, "ticker"]
                        raw_val = df_group.loc[idx, metric]
                        records.append((
                            ticker,
                            group,
                            metric,
                            float(raw_val) if (raw_val is not None and not pd.isna(raw_val)) else None,
                            float(pct),
                            int(year)
                        ))

        cursor.executemany("""
            INSERT INTO peer_percentiles (company_id, peer_group_name, metric, value, percentile_rank, year)
            VALUES (?, ?, ?, ?, ?, ?);
        """, records)
        conn.commit()
        conn.close()

    def generate_peer_comparison_report(self):
        """Generates output/peer_comparison.xlsx with 11 sheets."""
        conn = sqlite3.connect(self.db_path)

        latest_year_query = """
            SELECT 
                fr.*,
                c.name as company_name,
                c.sector_name as broad_sector,
                p.sales,
                p.net_income as net_profit,
                p.shares_outstanding
            FROM financial_ratios fr
            JOIN companies c ON fr.ticker = c.ticker
            LEFT JOIN profitandloss p ON fr.ticker = p.ticker AND fr.year = p.year
        """
        df = pd.read_sql_query(latest_year_query, conn)
        conn.close()

        df_sorted = df.sort_values(["ticker", "year"])
        latest_df = df_sorted.groupby("ticker").last().reset_index()

        engine_derived = self.compute_derived_for_report(latest_df)

        peer_dict = self.peer_df.set_index("ticker")["group_name"].to_dict()
        benchmark_dict = self.peer_df.set_index("ticker")["is_benchmark"].to_dict()

        engine_derived["peer_group"] = engine_derived["ticker"].map(peer_dict)
        engine_derived["is_benchmark"] = engine_derived["ticker"].map(benchmark_dict).fillna(0)

        kpi_columns = [
            "pe_ratio", "pb_ratio", "roe", "net_profit_margin_pct",
            "operating_profit_margin_pct", "return_on_capital_employed_pct", "return_on_assets_pct",
            "debt_to_equity", "interest_coverage", "asset_turnover",
            "free_cash_flow_cr", "cfo_quality_score", "fcf_conversion_rate_pct",
            "dividend_payout_ratio_pct", "dividend_yield", "revenue_cagr_5yr",
            "pat_cagr_5yr", "eps_cagr_5yr", "sales", "net_profit"
        ]

        lower_is_better_cols = {"pe_ratio", "pb_ratio", "debt_to_equity"}

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        groups = self.peer_df["group_name"].unique()

        for group in groups:
            group_df = engine_derived[engine_derived["peer_group"] == group].copy()
            if group_df.empty:
                continue

            group_df = group_df.sort_values("ticker")

            n = len(group_df)
            for col in kpi_columns:
                if col == "interest_coverage":
                    vals_to_rank = group_df.apply(
                        lambda r: float('inf') if r["icr_label"] == "Debt Free" else (r["interest_coverage"] or 0.0), axis=1
                    )
                else:
                    vals_to_rank = group_df[col].fillna(0.0)

                ranks = vals_to_rank.rank(method='min')
                if n <= 1:
                    pct_ranks = pd.Series(1.0, index=group_df.index)
                else:
                    pct_ranks = (ranks - 1.0) / (n - 1.0)

                if col in lower_is_better_cols:
                    pct_ranks = 1.0 - pct_ranks

                group_df[f"{col}_pct_rank"] = pct_ranks

            ws = wb.create_sheet(title=group[:30])

            ws.append([f"Peer Group Comparison Report: {group}"])
            ws.append(["Benchmark company is highlighted in GOLD."])
            ws.append(["Percentile ranks: Green (>=75th), Yellow (25th-75th), Red (<=25th)."])
            ws.append([])

            headers = ["Ticker", "Company Name", "Broad Sector"]
            for col in kpi_columns:
                headers.extend([f"{col} (Value)", f"{col} (% Rank)"])
            ws.append(headers)

            header_row_idx = 5
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=header_row_idx, column=col_idx)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = BORDER_THIN

            for _, row in group_df.iterrows():
                row_data = [row["ticker"], row["company_name"], row["broad_sector"]]
                for col in kpi_columns:
                    val = row[col]
                    pct = row[f"{col}_pct_rank"]

                    val_str = "-" if (val is None or pd.isna(val)) else round(val, 2)
                    pct_val = round(pct, 2)

                    row_data.extend([val_str, pct_val])
                ws.append(row_data)

                row_idx = ws.max_row
                is_bm = row["is_benchmark"] == 1

                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = BORDER_THIN
                    if is_bm:
                        cell.fill = BENCHMARK_GOLD

                for c_idx in range(len(kpi_columns)):
                    pct_col_idx = 5 + (c_idx * 2)
                    cell = ws.cell(row=row_idx, column=pct_col_idx)
                    pct_val = cell.value

                    if pct_val != "-":
                        if pct_val >= 0.75:
                            cell.fill = PERCENTILE_GREEN
                            cell.font = Font(color="006100", bold=True)
                        elif pct_val <= 0.25:
                            cell.fill = PERCENTILE_RED
                            cell.font = Font(color="9C0006", bold=True)
                        else:
                            cell.fill = PERCENTILE_YELLOW
                            cell.font = Font(color="9C6500", bold=True)

            median_row_data = ["Median", "", ""]
            for col in kpi_columns:
                med_val = group_df[col].dropna().median()
                med_val_str = "-" if pd.isna(med_val) else round(med_val, 2)
                med_pct = group_df[f"{col}_pct_rank"].dropna().median()
                med_pct_val = "-" if pd.isna(med_pct) else round(med_pct, 2)

                median_row_data.extend([med_val_str, med_pct_val])
            ws.append(median_row_data)

            med_row_idx = ws.max_row
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=med_row_idx, column=col_idx)
                cell.font = Font(bold=True)
                cell.border = BORDER_THIN
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

            ws.row_dimensions[header_row_idx].height = 28
            for col in range(1, len(headers) + 1):
                col_letter = openpyxl.utils.get_column_letter(col)
                max_len = max(len(str(ws.cell(row=r, column=col).value or '')) for r in range(header_row_idx, ws.max_row + 1))
                ws.column_dimensions[col_letter].width = max(max_len + 2, 10)

            ws.views.sheetView[0].showGridLines = True

        Path("output").mkdir(parents=True, exist_ok=True)
        out_path = "output/peer_comparison.xlsx"
        wb.save(out_path)

    def compute_derived_for_report(self, latest_df: pd.DataFrame) -> pd.DataFrame:
        """Helper to calculate derived columns for the latest year data (dividend_yield and market_cap)."""
        df = latest_df.copy()

        conn = sqlite3.connect(self.db_path)
        price_query = "SELECT ticker, strftime('%Y', date) as price_year, AVG(close) as avg_close FROM stock_prices GROUP BY ticker, price_year"
        df_prices = pd.read_sql_query(price_query, conn)
        df_prices["price_year"] = df_prices["price_year"].astype(int)
        prices_dict = df_prices.set_index(["ticker", "price_year"])["avg_close"].to_dict()

        fallback_query = "SELECT ticker, AVG(close) as overall_avg_close FROM stock_prices GROUP BY ticker"
        df_fallback = pd.read_sql_query(fallback_query, conn)
        fallback_prices = df_fallback.set_index("ticker")["overall_avg_close"].to_dict()

        corp_query = "SELECT ticker, date, value FROM corporate_actions WHERE action_type = 'Dividend'"
        df_corp = pd.read_sql_query(corp_query, conn)
        if not df_corp.empty:
            df_corp["year"] = pd.to_datetime(df_corp["date"]).dt.year
            dividend_dict = df_corp.groupby(["ticker", "year"])["value"].sum().to_dict()
        else:
            dividend_dict = {}

        conn.close()

        df["close_price"] = df.apply(
            lambda r: prices_dict.get((r.ticker, int(r.year)), fallback_prices.get(r.ticker, 0.0)), axis=1
        )
        df["dps"] = df.apply(
            lambda r: dividend_dict.get((r.ticker, int(r.year)), 0.0), axis=1
        )
        df["dividend_yield"] = df.apply(
            lambda r: (r.dps / r.close_price * 100.0) if r.close_price > 0.0 else 0.0, axis=1
        )
        df["market_cap"] = df.apply(
            lambda r: (r.close_price * r.shares_outstanding) if (r.close_price is not None and r.shares_outstanding is not None) else None, axis=1
        )

        return df

    def generate_radar_charts(self):
        """Generates radar charts overlaying company metrics vs peer averages, or standalone bar charts."""
        reports_dir = Path("reports/radar_charts")
        reports_dir.mkdir(parents=True, exist_ok=True)

        conn = sqlite3.connect(self.db_path)
        query = """
            SELECT 
                fr.*,
                c.name as company_name,
                c.sector_name as broad_sector,
                p.sales,
                p.net_income as net_profit,
                p.shares_outstanding
            FROM financial_ratios fr
            JOIN companies c ON fr.ticker = c.ticker
            LEFT JOIN profitandloss p ON fr.ticker = p.ticker AND fr.year = p.year
        """
        df = pd.read_sql_query(query, conn)
        conn.close()

        df_sorted = df.sort_values(["ticker", "year"])
        latest_df = df_sorted.groupby("ticker").last().reset_index()

        latest_df = self.compute_derived_for_report(latest_df)

        peer_dict = self.peer_df.set_index("ticker")["group_name"].to_dict()
        latest_df["peer_group"] = latest_df["ticker"].map(peer_dict)

        radar_metrics = [
            "roe", "return_on_capital_employed_pct", "net_profit_margin_pct",
            "debt_to_equity", "free_cash_flow_cr", "pat_cagr_5yr",
            "revenue_cagr_5yr", "composite_quality_score"
        ]
        labels = [
            "ROE", "ROCE", "NPM", "D/E (Inverted)", "FCF", "PAT CAGR 5Yr", "Rev CAGR 5Yr", "Composite Score"
        ]

        mapped_df = latest_df[latest_df["peer_group"].notna()].copy()
        unmapped_df = latest_df[latest_df["peer_group"].isna()].copy()

        n_mapped = len(mapped_df)
        if n_mapped > 0:
            groups = mapped_df["peer_group"].unique()
            for group in groups:
                g_mask = mapped_df["peer_group"] == group
                g_df = mapped_df[g_mask]
                g_len = len(g_df)

                for metric in radar_metrics:
                    vals = g_df[metric].fillna(0.0)
                    ranks = vals.rank(method='min')

                    if g_len <= 1:
                        scores = pd.Series(100.0, index=vals.index)
                    else:
                        scores = (ranks - 1.0) / (g_len - 1.0) * 100.0

                    if metric == "debt_to_equity":
                        scores = 100.0 - scores

                    mapped_df.loc[g_mask, f"{metric}_score"] = scores

            for idx, company in mapped_df.iterrows():
                ticker = company["ticker"]
                name = company["company_name"]
                group = company["peer_group"]

                comp_scores = [company[f"{m}_score"] for m in radar_metrics]

                group_mask = mapped_df["peer_group"] == group
                group_companies = mapped_df[group_mask]
                avg_scores = [group_companies[f"{m}_score"].mean() for m in radar_metrics]

                self._draw_radar(
                    ticker=ticker,
                    company_name=name,
                    group_name=group,
                    labels=labels,
                    values=comp_scores,
                    avg_values=avg_scores,
                    save_path=reports_dir / f"{ticker}_radar.png"
                )

        nifty100_avg_score = latest_df["composite_quality_score"].mean()

        for idx, company in unmapped_df.iterrows():
            ticker = company["ticker"]
            name = company["company_name"]
            comp_score = company["composite_quality_score"]

            self._draw_standalone_chart(
                ticker=ticker,
                company_name=name,
                value=comp_score,
                avg_value=nifty100_avg_score,
                save_path=reports_dir / f"{ticker}_radar.png"
            )

    def _draw_radar(self, ticker: str, company_name: str, group_name: str, labels: List[str], values: List[float], avg_values: List[float], save_path: Path):
        num_vars = len(labels)
        angles = np.linspace(0, 2 * np.pi, num_vars, endpoint=False).tolist()

        draw_values = values + values[:1]
        draw_avg = avg_values + avg_values[:1]
        draw_angles = angles + angles[:1]

        fig, ax = plt.subplots(figsize=(6, 6), subplot_kw=dict(polar=True))

        plt.xticks(angles, labels, color="grey", size=8)

        ax.set_rlabel_position(0)  # type: ignore
        plt.yticks([25, 50, 75, 100], ["25", "50", "75", "100"], color="grey", size=7)
        plt.ylim(0, 100)

        ax.plot(draw_angles, draw_values, linewidth=2, linestyle='solid', label=company_name, color='#1F4E78')
        ax.fill(draw_angles, draw_values, color='#1F4E78', alpha=0.25)

        ax.plot(draw_angles, draw_avg, linewidth=2, linestyle='dashed', label=f"{group_name} Avg", color='#C00000')

        plt.title(f"Radar Chart: {company_name}\n(vs {group_name} Peers)", size=10, color='#1F4E78', y=1.1, weight='bold', ha='center')
        plt.legend(loc='upper right', bbox_to_anchor=(0.1, 0.1), fontsize=7)

        plt.tight_layout()
        plt.savefig(save_path, dpi=150)
        plt.close()

    def _draw_standalone_chart(self, ticker: str, company_name: str, value: float, avg_value: float, save_path: Path):
        fig, ax = plt.subplots(figsize=(5, 5))
        display_names = [
            company_name[:15] + "..." if len(company_name) > 15 else company_name,
            "Nifty 100 Avg",
        ]
        values = [value, avg_value]

        bars = ax.bar(display_names, values, color=["#1F4E78", "#7F7F7F"], width=0.4)

        ax.set_ylabel("Composite Quality Score (0-100)")
        ax.set_ylim(0, 100)

        for bar in bars:
            h = bar.get_height()
            ax.text(
                bar.get_x() + bar.get_width() / 2.0,
                h + 1.5,
                f"{h:.2f}",
                ha="center",
                va="bottom",
                weight="bold",
                size=9,
            )

        plt.title(f"Composite Score: {company_name}\n(No peer group assigned)", size=10, color='#1F4E78', weight='bold', y=1.05)
        plt.tight_layout()
        plt.savefig(save_path, dpi=150)
        plt.close()

def main():
    analyzer = PeerAnalyzer()
    analyzer.calculate_percentiles_and_save()
    analyzer.generate_peer_comparison_report()
    analyzer.generate_radar_charts()

if __name__ == '__main__':
    main()
