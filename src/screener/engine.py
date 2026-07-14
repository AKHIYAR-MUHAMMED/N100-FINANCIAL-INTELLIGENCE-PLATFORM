import os
import sqlite3
from pathlib import Path
from typing import Any, Dict

import openpyxl
import pandas as pd
import yaml
from openpyxl.styles import Alignment, Font, PatternFill

from src.analytics.cagr import calculate_cagr

# Define standard colors for Excel formatting
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Pastel Green
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # Pastel Red
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

class ScreenerEngine:
    def __init__(self, db_path: str = "data/db/nifty100.db", config_path: str = "config/screener_config.yaml"):
        self.db_path = db_path
        self.config_path = config_path
        self.config = self.load_config()

    def load_config(self) -> Dict[str, Any]:
        """Loads screener configuration from YAML file."""
        if not os.path.exists(self.config_path):
            raise FileNotFoundError(f"Config file not found at: {self.config_path}")
        with open(self.config_path, "r", encoding="utf-8") as f:
            return yaml.safe_load(f)

    def load_and_prepare_data(self) -> pd.DataFrame:
        """Fetches data from SQLite, computes derived columns and sector scores."""
        conn = sqlite3.connect(self.db_path)

        # 1. Fetch ratios and financials combined
        query = """
            SELECT 
                fr.*,
                c.name as company_name,
                c.sector_name as broad_sector,
                c.industry,
                p.sales,
                p.net_income as net_profit,
                p.shares_outstanding
            FROM financial_ratios fr
            JOIN companies c ON fr.ticker = c.ticker
            LEFT JOIN profitandloss p ON fr.ticker = p.ticker AND fr.year = p.year
            ORDER BY fr.ticker, fr.year
        """
        df = pd.read_sql_query(query, conn)

        # 2. Fetch stock prices and corporate actions
        price_query = """
            SELECT
                ticker,
                strftime('%Y', date) as price_year,
                AVG(close) as avg_close
            FROM stock_prices
            GROUP BY ticker, price_year
        """
        df_prices = pd.read_sql_query(price_query, conn)
        df_prices["price_year"] = df_prices["price_year"].astype(int)
        prices_dict = df_prices.set_index(["ticker", "price_year"])["avg_close"].to_dict()

        fallback_query = "SELECT ticker, AVG(close) as overall_avg_close FROM stock_prices GROUP BY ticker"
        df_fallback = pd.read_sql_query(fallback_query, conn)
        fallback_prices = df_fallback.set_index("ticker")["overall_avg_close"].to_dict()

        corp_query = """
            SELECT ticker, date, value 
            FROM corporate_actions 
            WHERE action_type = 'Dividend'
        """
        df_corp = pd.read_sql_query(corp_query, conn)
        if not df_corp.empty:
            df_corp["year"] = pd.to_datetime(df_corp["date"]).dt.year
            dividend_dict = df_corp.groupby(["ticker", "year"])["value"].sum().to_dict()
        else:
            dividend_dict = {}

        conn.close()

        # 3. Calculate derived columns
        df["close_price"] = df.apply(
            lambda r: prices_dict.get((r.ticker, int(r.year)), fallback_prices.get(r.ticker, 0.0)), axis=1
        )
        df["dps"] = df.apply(
            lambda r: dividend_dict.get((r.ticker, int(r.year)), 0.0), axis=1
        )
        df["dividend_yield"] = df.apply(
            lambda r: (r.dps / r.close_price * 100.0) if r.close_price > 0.0 else 0.0, axis=1
        )
        df["market_cap"] = df.apply(
            lambda r: (r.close_price * r.shares_outstanding) if (r.close_price is not None and r.shares_outstanding is not None) else None, axis=1
        )
        df["cfo_pat_ratio"] = df.apply(
            lambda r: (r.cash_from_operations_cr / r.net_profit) if (r.net_profit is not None and r.net_profit > 0.0 and r.cash_from_operations_cr is not None) else 0.0, axis=1
        )
        df["fcf_positive_flag"] = df["free_cash_flow_cr"].apply(
            lambda val: 1.0 if (val is not None and val > 0.0) else 0.0
        )

        fcf_dict = df.set_index(["ticker", "year"])["free_cash_flow_cr"].to_dict()

        def get_fcf_cagr(row):
            ticker = row["ticker"]
            year = int(row["year"])
            fcf_latest = row["free_cash_flow_cr"]
            fcf_start = fcf_dict.get((ticker, year - 5))

            if fcf_latest is None or fcf_start is None:
                return 0.0

            val, flag = calculate_cagr(fcf_start, fcf_latest, 5)
            return val if val is not None else 0.0

        df["fcf_cagr"] = df.apply(get_fcf_cagr, axis=1)

        score_cols = [
            "roe", "return_on_capital_employed_pct", "net_profit_margin_pct",
            "fcf_cagr", "cfo_pat_ratio", "fcf_positive_flag",
            "revenue_cagr_5yr", "pat_cagr_5yr", "debt_to_equity", "interest_coverage"
        ]
        df["interest_coverage_scored"] = df.apply(
            lambda r: float('inf') if r.icr_label == "Debt Free" else (r.interest_coverage or 0.0), axis=1
        )

        for col in score_cols:
            if col != "interest_coverage":
                df[col] = df[col].fillna(0.0)

        df_scored = self.compute_sector_scores(df)
        self.save_composite_scores_to_db(df_scored)

        return df_scored

    def compute_sector_scores(self, df: pd.DataFrame) -> pd.DataFrame:
        """Computes winsorised and sector-relative composite quality score."""
        df_scores = df.copy()

        score_mapping = {
            "roe": ("roe_score", True, False),
            "return_on_capital_employed_pct": ("roce_score", True, False),
            "net_profit_margin_pct": ("npm_score", True, False),
            "fcf_cagr": ("fcf_cagr_score", True, False),
            "cfo_pat_ratio": ("cfo_pat_score", True, False),
            "fcf_positive_flag": ("fcf_flag_score", True, False),
            "revenue_cagr_5yr": ("rev_cagr_score", True, False),
            "pat_cagr_5yr": ("pat_cagr_score", True, False),
            "debt_to_equity": ("de_score", False, False),
            "interest_coverage_scored": ("icr_score", True, True)
        }

        for score_col in [v[0] for v in score_mapping.values()]:
            df_scores[score_col] = 0.0

        sectors = df_scores["broad_sector"].unique()

        for sector in sectors:
            sector_mask = df_scores["broad_sector"] == sector
            sector_df = df_scores[sector_mask]

            if sector_df.empty:
                continue

            for col_name, (score_col, higher_is_better, handle_inf_icr) in score_mapping.items():
                vals = sector_df[col_name].copy()

                if handle_inf_icr:
                    is_inf = (vals == float('inf'))
                    finite_mask = ~is_inf & vals.notna()
                    finite_vals = vals[finite_mask]
                else:
                    finite_mask = vals.notna()
                    finite_vals = vals[finite_mask]

                if finite_vals.empty:
                    final_scores = pd.Series(0.0, index=vals.index)
                    if handle_inf_icr:
                        final_scores[is_inf] = 100.0
                    df_scores.loc[sector_mask, score_col] = final_scores
                    continue

                p10 = finite_vals.quantile(0.1)
                p90 = finite_vals.quantile(0.9)

                if p90 == p10:
                    finite_scores = pd.Series(100.0, index=finite_vals.index)
                else:
                    capped = finite_vals.clip(lower=p10, upper=p90)
                    if higher_is_better:
                        finite_scores = (capped - p10) / (p90 - p10) * 100.0
                    else:
                        finite_scores = (p90 - capped) / (p90 - p10) * 100.0

                final_scores = pd.Series(0.0, index=vals.index)
                final_scores[finite_vals.index] = finite_scores
                if handle_inf_icr:
                    final_scores[is_inf] = 100.0

                df_scores.loc[sector_mask, score_col] = final_scores

        df_scores["composite_quality_score"] = (
            0.15 * df_scores["roe_score"] +
            0.10 * df_scores["roce_score"] +
            0.10 * df_scores["npm_score"] +
            0.15 * df_scores["fcf_cagr_score"] +
            0.10 * df_scores["cfo_pat_score"] +
            0.05 * df_scores["fcf_flag_score"] +
            0.10 * df_scores["rev_cagr_score"] +
            0.10 * df_scores["pat_cagr_score"] +
            0.10 * df_scores["de_score"] +
            0.05 * df_scores["icr_score"]
        )

        return df_scores

    def save_composite_scores_to_db(self, df_scores: pd.DataFrame):
        """Updates composite_quality_score column in SQLite financial_ratios table."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        data = [(row.composite_quality_score, row.ticker, int(row.year)) for row in df_scores.itertuples()]

        cursor.executemany(
            "UPDATE financial_ratios SET composite_quality_score = ? WHERE ticker = ? AND year = ?",
            data
        )
        conn.commit()
        conn.close()

    def get_latest_company_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """Extracts the latest year data for each company, computing de_declining flags."""
        df_sorted = df.sort_values(["ticker", "year"])
        df_sorted["prev_de"] = df_sorted.groupby("ticker")["debt_to_equity"].shift(1)
        df_sorted["de_declining"] = df_sorted["debt_to_equity"] < df_sorted["prev_de"]
        df_sorted["prev_sales"] = df_sorted.groupby("ticker")["sales"].shift(1)

        latest_df = df_sorted.groupby("ticker").last().reset_index()
        return latest_df

    def apply_preset_filters(self, latest_df: pd.DataFrame, preset_name: str) -> pd.DataFrame:
        """Filters companies based on preset criteria."""
        preset = self.config["presets"].get(preset_name)
        if not preset:
            raise ValueError(f"Unknown preset name: {preset_name}")

        filters = preset["filters"]
        filtered = latest_df.copy()

        for metric, criteria in filters.items():
            if metric == "de_declining":
                filtered = filtered[filtered["de_declining"]]
                continue

            if metric == "debt_to_equity":
                if "max" in criteria:
                    limit = criteria["max"]
                    filtered = filtered[(filtered["broad_sector"] == "Financials") | (filtered["debt_to_equity"] <= limit)]
                if "min" in criteria:
                    limit = criteria["min"]
                    filtered = filtered[(filtered["broad_sector"] == "Financials") | (filtered["debt_to_equity"] >= limit)]
                if "exact" in criteria:
                    limit = criteria["exact"]
                    filtered = filtered[(filtered["broad_sector"] == "Financials") | (filtered["debt_to_equity"] == limit)]
                continue

            if metric == "interest_coverage":
                if "min" in criteria:
                    limit = criteria["min"]
                    filtered = filtered[(filtered["icr_label"] == "Debt Free") | (filtered["interest_coverage"] >= limit)]
                if "max" in criteria:
                    limit = criteria["max"]
                    filtered = filtered[(filtered["icr_label"] == "Debt Free") | (filtered["interest_coverage"] <= limit)]
                continue

            if metric in filtered.columns:
                if "min" in criteria:
                    filtered = filtered[filtered[metric] >= criteria["min"]]
                if "max" in criteria:
                    filtered = filtered[filtered[metric] <= criteria["max"]]
                if "exact" in criteria:
                    filtered = filtered[filtered[metric] == criteria["exact"]]
            elif metric == "sales" and "sales" in filtered.columns:
                if "min" in criteria:
                    filtered = filtered[filtered["sales"] >= criteria["min"]]
            elif metric == "dividend_payout_ratio_pct" and "dividend_payout_ratio_pct" in filtered.columns:
                if "max" in criteria:
                    filtered = filtered[filtered["dividend_payout_ratio_pct"] <= criteria["max"]]
            elif metric == "revenue_cagr_3yr" and "revenue_cagr_3yr" in filtered.columns:
                if "min" in criteria:
                    filtered = filtered[filtered["revenue_cagr_3yr"] >= criteria["min"]]

        return filtered.sort_values("composite_quality_score", ascending=False)

    def generate_screener_output(self, df_all: pd.DataFrame):
        """Runs all 6 presets and generates output/screener_output.xlsx."""
        latest_df = self.get_latest_company_data(df_all)

        kpi_columns = [
            "pe_ratio", "pb_ratio", "roe", "net_profit_margin_pct",
            "operating_profit_margin_pct", "return_on_capital_employed_pct", "return_on_assets_pct",
            "debt_to_equity", "interest_coverage", "asset_turnover",
            "free_cash_flow_cr", "cfo_quality_score", "fcf_conversion_rate_pct",
            "dividend_payout_ratio_pct", "dividend_yield", "revenue_cagr_5yr",
            "pat_cagr_5yr", "eps_cagr_5yr", "sales", "net_profit"
        ]

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        preset_keys = [
            "quality_compounder", "value_pick", "growth_accelerator",
            "dividend_champion", "debt_free_blue_chip", "turnaround_watch"
        ]

        for key in preset_keys:
            preset_config = self.config["presets"][key]
            sheet_title = preset_config["name"]

            df_filtered = self.apply_preset_filters(latest_df, key)
            ws = wb.create_sheet(title=sheet_title)

            ws.append([f"Screener Preset: {sheet_title}"])
            ws.append([preset_config["description"]])
            ws.append([f"Companies found: {len(df_filtered)} (Sorted by Composite Score descending)"])
            ws.append([])

            headers = ["Ticker", "Company Name", "Broad Sector", "Composite Score"] + kpi_columns
            ws.append(headers)

            header_row_idx = 5
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=header_row_idx, column=col_idx)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            filters = preset_config["filters"]
            for idx, row in df_filtered.iterrows():
                row_data = [
                    row["ticker"],
                    row["company_name"],
                    row["broad_sector"],
                    round(row["composite_quality_score"], 2)
                ]
                for col in kpi_columns:
                    val = row[col]
                    if pd.isna(val):
                        row_data.append("-")
                    else:
                        row_data.append(round(val, 4) if isinstance(val, float) else val)
                ws.append(row_data)

                data_row_idx = ws.max_row

                for f_col, criteria in filters.items():
                    map_col = f_col
                    if f_col == "de_declining":
                        map_col = "debt_to_equity"
                    elif f_col == "sales":
                        map_col = "sales"
                    elif f_col == "dividend_payout_ratio_pct":
                        map_col = "dividend_payout_ratio_pct"
                    elif f_col == "revenue_cagr_3yr":
                        map_col = "revenue_cagr_5yr"

                    if map_col in kpi_columns:
                        kpi_col_idx = kpi_columns.index(map_col)
                        sheet_col_idx = 5 + kpi_col_idx
                        cell = ws.cell(row=data_row_idx, column=sheet_col_idx)

                        val = row[map_col]
                        is_financial = row["broad_sector"] == "Financials"
                        is_debt_free = row["icr_label"] == "Debt Free"

                        passed = True
                        if f_col == "de_declining":
                            passed = bool(row["de_declining"])
                        elif f_col == "debt_to_equity" and is_financial:
                            passed = True
                        elif f_col == "interest_coverage" and is_debt_free:
                            passed = True
                        else:
                            if pd.isna(val) or val == "-":
                                passed = False
                            else:
                                if "min" in criteria:
                                    passed = passed and (val >= criteria["min"])
                                if "max" in criteria:
                                    passed = passed and (val <= criteria["max"])
                                if "exact" in criteria:
                                    passed = passed and (val == criteria["exact"])

                        cell.fill = GREEN_FILL if passed else RED_FILL
                        cell.font = Font(color="006100" if passed else "9C0006")

            ws.row_dimensions[header_row_idx].height = 28
            for col_idx in range(1, len(headers) + 1):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                max_len = max(
                    len(str(ws.cell(row=r, column=col_idx).value or ""))
                    for r in range(header_row_idx, ws.max_row + 1)
                )
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

            ws.views.sheetView[0].showGridLines = True

        Path("output").mkdir(parents=True, exist_ok=True)
        out_path = "output/screener_output.xlsx"
        wb.save(out_path)

def main():
    engine = ScreenerEngine()
    df_all = engine.load_and_prepare_data()
    engine.generate_screener_output(df_all)

if __name__ == '__main__':
    main()
